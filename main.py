import kivy
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.image import Image
from kivy.config import Config
from odf.text import P
from tkinter import Tk
from tkinter.filedialog import asksaveasfilename
from kivy.graphics import Color, Rectangle
from function_file import date_du_jour, heure_du_jour
#from function_vocale import speak, enregistrer_vocale
from kivy.uix.spinner import Spinner, SpinnerOption
from kivy.uix.checkbox import CheckBox
from kivy.core.window import Window
from kivy.uix.popup import Popup
from PIL import ImageGrab
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
import os
from kivy.uix.scrollview import ScrollView
from kivy.uix.behaviors import ButtonBehavior
from kivy.lang import Builder



class MyBoxLayout(BoxLayout):
    # Initialisation
    def __init__(self, **kwargs):
                
        # Appelle le constructeur de GridLayout
        super(MyBoxLayout, self).__init__(**kwargs)
        self.orientation = "vertical"
        self.padding = [10,0 ,10, 0]        
        self.spacing = 10    

        self.text_inputs = []
        

        # Capturez les événements de clavier
        Window.bind(on_key_down=self.on_key_down)

        # Créer une instance de GridLayout
        self.my_gridlayout = GridLayout(cols=3, spacing=0, padding=0, size_hint=(1, None), height=50)

        # Ajouter des widgets à la GridLayout
            # Ajoute un bouton
        self.my_gridlayout1 = GridLayout(rows=1, spacing=0, padding=0, size_hint=(None, None), size=(250, 54))       
        # Ajoutez un widget Image
        self.my_image = Image(source='logo62.jpg')
        
        self.my_gridlayout1.add_widget(self.my_image)
        
        self.my_gridlayout.add_widget(self.my_gridlayout1)

        self.my_gridlayou_agent = GridLayout(rows=1, spacing=0, padding=0, size_hint=(0.25, 1))


        self.options = ['Prenom1', 'Prenom2', 'Prenom3', 'Prenom4', 'Prenom5']
        self.my_spinner_agent = Spinner(
            text='Nom Référent',  # Texte initial
            values=self.options, color='#FFFFFF',background_color=(0,1,1,1),
             option_cls=self.custom_spinner_option, font_size=25)  # Liste d'options
            #size_hint=(None, None))
            #size=(150, 44))  # Taille du menu déroulant)
        def on_spinner_select(instance, value):
            # La fonction est appelée lorsque l'utilisateur sélectionne une option
            print(f'Option sélectionnée : {value}')

        self.my_spinner_agent.bind(text=on_spinner_select)

        self.my_gridlayou_agent.add_widget(self.my_spinner_agent)
        
        self.my_gridlayout.add_widget(self.my_gridlayou_agent)

            # Ajoute un bouton
        self.my_gridlayout2 = GridLayout(rows=1, spacing=0, padding=0, size_hint=(0.3, 1))
        
        self.options = ['Demandeur1', 'Demandeur2', 'Demandeur3', 'Demandeur4', 'Demandeur5']
        self.my_spinner_cellule = Spinner(
            text='Demandeur',  # Texte initial
            values=self.options, color='#FFFFFF',background_color=(1,1,1,0.3), option_cls=self.custom_spinner_option, font_size=25)  # Liste d'options
            #size_hint=(None, None))
            #size=(150, 44))  # Taille du menu déroulant)
        def on_spinner_select(instance, value):
            # La fonction est appelée lorsque l'utilisateur sélectionne une option
            print(f'Option sélectionnée : {value}')

        self.my_spinner_cellule.bind(text=on_spinner_select)

        self.my_gridlayout2.add_widget(self.my_spinner_cellule)

        self.my_gridlayout.add_widget(self.my_gridlayout2)
        
        
        # Ajouter la GridLayout à la BoxLayout
        self.add_widget(self.my_gridlayout)

        #-----------------------------------------------------------------------------------------------------------------------------------------

        self.compteur = 0
            # Ajoute un bouton
        self.my_gridlayout3b = GridLayout(cols=3, spacing=0, padding=0, size_hint=(1, 1))

        
        
        self.my_gridlayout2 = GridLayout(rows=1, spacing=0, padding=0, size_hint=(0.6, 1))
            # Ajoute un bouton
        self.my_gridlayout3 = GridLayout(rows=2, spacing=0, padding=0, size_hint=(0.1, 1))
        self.my_button_envoyer3 = Button(text="Date", font_size=15, background_color=(0.1,0.56,0.9,1))
        self.my_button_envoyer4 = Button(text="Heure", font_size=15, background_color=(0.43,0.88,1.2,1))
        self.my_gridlayout3.add_widget(self.my_button_envoyer3)
        self.my_gridlayout3.add_widget(self.my_button_envoyer4)

        self.my_gridlayout2.add_widget(self.my_gridlayout3)
        
        
            # Ajoute un bouton
        self.my_gridlayout4 = GridLayout(rows=2, spacing=0, padding=0, size_hint=(0.1, 1))
        self.my_button_date= Button(text=f"{date_du_jour()}", font_size=15, background_color=(0.1,0.56,0.9,1))
        self.my_button_heure = Button(text=f"{heure_du_jour()}", font_size=15, background_color=(0.43,0.88,1.2,1))
        self.my_gridlayout4.add_widget(self.my_button_date)
        self.my_gridlayout4.add_widget(self.my_button_heure)

        self.my_gridlayout2.add_widget(self.my_gridlayout4)
        #------------------------------------

        self.my_gridlayout3b.add_widget(self.my_gridlayout2)
        
        self.my_gridlayout_compteur = Button(text=f"{self.compteur} tache(s)", font_size=20, color='#FFFFFF',background_color=(0,1,1,1))

        self.my_gridlayout3b.add_widget(self.my_gridlayout_compteur)

        
        
        self.my_gridlayout2 = GridLayout(rows=1, spacing=0, padding=0, size_hint=(0.6, 1))
            # Ajoute un bouton
        self.my_gridlayout3 = GridLayout(rows=2, spacing=0, padding=0, size_hint=(0.1, 1))
        self.my_button_heure_debut = Button(text="Début", font_size=15, background_color=(0.43,0.88,1.2,1))
        self.my_button_heure_debut_value = Button(text=f"{heure_du_jour()}", font_size=15, background_color=(0.43,0.88,1.2,1))
         
        self.my_button_heure_debut.bind(on_press=self.reset_heure_debut)
        self.my_gridlayout3.add_widget(self.my_button_heure_debut)
        self.my_gridlayout3.add_widget(self.my_button_heure_debut_value)

        self.my_gridlayout2.add_widget(self.my_gridlayout3)
        
        
            # Ajoute un bouton
        self.my_gridlayout4 = GridLayout(rows=2, spacing=0, padding=0, size_hint=(0.1, 1))
        self.my_button_heure_fin= Button(text="Fin", font_size=15, background_color=(0.1,0.56,0.9,1))
        self.my_button_heure_fin_value = Button(text=f"{heure_du_jour()}", font_size=15, background_color=(0.1,0.56,0.9,1)) 

        self.my_button_heure_fin.bind(on_press=self.reset_heure_fin)
        self.my_gridlayout4.add_widget(self.my_button_heure_fin)
        self.my_gridlayout4.add_widget(self.my_button_heure_fin_value) 

        self.my_gridlayout2.add_widget(self.my_gridlayout4)
        #------------------------------------

        self.my_gridlayout3b.add_widget(self.my_gridlayout2)
        
        self.add_widget(self.my_gridlayout3b)

        #1111-------------------------------------------------------------------------------------------------------------------------------------

            # Ajoute un bouton
            

        self.my_gridlayout3b = GridLayout(cols=1, spacing=10, padding=5, size_hint=(1, 3))



        self.my_gridlayout3b1 = GridLayout(rows=4, spacing=0, padding=0, size_hint=(0.25, 0.25))

        self.my_gridlayout3ba = GridLayout(cols=2, spacing=0, padding=0, size_hint=(0.25, 1))  
        self.my_gridlayout3bammm = GridLayout(cols=1, spacing=0, padding=0, size_hint=(None, 1), width=150)       
        self.my_button_type = Button(text="Type de contact", font_size=19, background_color=(0.1,0.56,0.9,1))
        
        
        self.my_gridlayout3bannn = GridLayout(cols=1, spacing=0, padding=0, size_hint=(1, 1)) 
        self.options = ['MAIL', 'PAO', 'CLUB UTILISATEUR', 'COMITE RP', 'WEBINAIRE']
        self.my_spinner_type = Spinner(
            text='Sélectionnez de contact',  # Texte initial
            values=self.options, option_cls=self.custom_spinner_option, color='#00000',background_color=(255,255,255,1), font_size=19)  # Liste d'options
            #size_hint=(None, None))
            #size=(150, 44))  # Taille du menu déroulant)
        def on_spinner_select(instance, value):
            # La fonction est appelée lorsque l'utilisateur sélectionne une option
            print(f'Option sélectionnée : {value}')

        self.my_spinner_type.bind(text=on_spinner_select)
        
        self.my_gridlayout3bammm.add_widget(self.my_button_type)
        self.my_gridlayout3ba.add_widget(self.my_gridlayout3bammm)
        self.my_gridlayout3bannn.add_widget(self.my_spinner_type)
        self.my_gridlayout3ba.add_widget(self.my_gridlayout3bannn)

        self.my_gridlayout3b1.add_widget(self.my_gridlayout3ba)

        self.my_gridlayout3ba1 = GridLayout(cols=2, spacing=0, padding=0, size_hint=(0.25, 1))   
        
        self.my_gridlayout3bayyy = GridLayout(cols=1, spacing=0, padding=0, size_hint=(None, 1), width=150)              
        self.my_button_nom = Button(text="Sujet n°1", font_size=19, background_color=(0.1,0.56,0.9,1))
        
        #self.my_button_nom.bind(on_press=self.enregistrer_vocale_nom)

        self.my_gridlayout3bauuu = GridLayout(cols=1, spacing=0, padding=0, size_hint=(1, 1)) 
        self.my_text_input_1 = TextInput(multiline=True, font_size=19)
        self.my_gridlayout3bayyy.add_widget(self.my_button_nom)
        self.my_gridlayout3ba1.add_widget(self.my_gridlayout3bayyy)

        self.my_gridlayout3ba1.add_widget(self.my_text_input_1)       
        self.text_inputs.append(self.my_text_input_1)

        self.my_gridlayout3b1.add_widget(self.my_gridlayout3ba1)

        self.my_gridlayout3ba1 = GridLayout(cols=2, spacing=0, padding=0, size_hint=(0.25, 1))     
        
        self.my_gridlayout3bazzz = GridLayout(cols=1, spacing=0, padding=0, size_hint=(None, 1), width=150)      
        self.my_button_envoyer3b1 = Button(text="Sujet n°2", font_size=19, background_color=(0.1,0.56,0.9,1))
        self.my_text_input_2 = TextInput(multiline=True, font_size=19)
        self.my_gridlayout3bazzz.add_widget(self.my_button_envoyer3b1)
        self.my_gridlayout3ba1.add_widget(self.my_gridlayout3bazzz)
        self.my_gridlayout3ba1.add_widget(self.my_text_input_2)      
        self.text_inputs.append(self.my_text_input_2)

        self.my_gridlayout3b1.add_widget(self.my_gridlayout3ba1)

        self.my_gridlayout3ba1 = GridLayout(cols=2, spacing=0, padding=0, size_hint=(0.25, 1))   
        
        self.my_gridlayout3baqqq = GridLayout(cols=1, spacing=0, padding=0, size_hint=(None, 1), width=150)        
        self.my_button_territoire = Button(text="Type de sujet", font_size=19, background_color=(0.1,0.56,0.9,1))

        self.options = ['Prestations', 'Outils données', 'Courriers']
        self.my_spinner_territoire = Spinner(
            text='Sélectionnez un sujet', color='#00000',background_color=(255,255,255,1),  # Texte initial
            values=self.options, option_cls=self.custom_spinner_option, font_size=19)  # Liste d'options
            #size_hint=(None, None))
            #size=(150, 44))  # Taille du menu déroulant)
        def on_spinner_select(instance, value):
            # La fonction est appelée lorsque l'utilisateur sélectionne une option
            print(f'Option sélectionnée : {value}')

        self.my_spinner_territoire.bind(text=on_spinner_select)

        self.my_gridlayout3baqqq.add_widget(self.my_button_territoire)
        self.my_gridlayout3ba1.add_widget(self.my_gridlayout3baqqq)
        self.my_gridlayout3ba1.add_widget(self.my_spinner_territoire)

        self.my_gridlayout3b1.add_widget(self.my_gridlayout3ba1)

        self.my_gridlayout3b.add_widget(self.my_gridlayout3b1)



        #-------------------------------------------------------------------------------
        


        self.add_widget(self.my_gridlayout3b)

        #2222-----------------------------------------------------------------------------------------------------------------------------------

            # Ajoute un bouton
            

        self.my_gridlayout3b = GridLayout(cols=1, spacing=10, padding=5, size_hint=(1, 3))



        self.my_gridlayout3b1 = GridLayout(rows=4, spacing=0, padding=0, size_hint=(0.25, 0.25))

        
        # Créez des listes pour les options de chaque Spinner
        categories = ['Référenciel Fonctionnel', "DEMANDE D'INFORMATIONS DIVERSES", 'EXPLICATION DE DÉCISION',
                       "DEMANDE D'ENVOI DE DOCUMENTS DIVERS" , "RECTIFICATIF DE DÉCISION"]


        self.my_gridlayout3ba = GridLayout(cols=2, spacing=0, padding=0, size_hint=(0.25, 1.5))  
        
        self.my_gridlayout3baxxx = GridLayout(cols=1, spacing=0, padding=0, size_hint=(None, 1), width=150)        
        self.my_button_motif1 = Button(text="Réf 1", font_size=19, background_color=(0.43,0.88,1.2,1))
        self.my_button_motif1.bind(on_press=self.reset_spinner_parent)

        # Créez les Spinners
        self.spinner_parent = Spinner(
            text='Document',
            values=categories, option_cls=self.custom_spinner_option, color='#00000',background_color=(255,255,255,1), font_size=19
        )

        self.my_gridlayout3baxxx.add_widget(self.my_button_motif1)
        self.my_gridlayout3ba.add_widget(self.my_gridlayout3baxxx)
        self.my_gridlayout3ba.add_widget(self.spinner_parent)

        self.my_gridlayout3b1.add_widget(self.my_gridlayout3ba)

        self.my_gridlayout3ba1 = GridLayout(cols=2, spacing=0, padding=0, size_hint=(0.25, 1.5)) 
        
        self.my_gridlayout3bawww = GridLayout(cols=1, spacing=0, padding=0, size_hint=(None, 1), width=150)         
        self.my_button_motif2 = Button(text="Réf 2", font_size=19, background_color=(0.43,0.88,1.2,1))        
        self.my_button_motif2.bind(on_press=self.reset_spinner_child)

        
        # Spinner enfant
      
        self.spinner_child = Spinner(
            text='Version document',
            values=[], option_cls=self.custom_spinner_option, color='#00000',background_color=(255,255,255,1), font_size=19
        )

        self.my_gridlayout3bawww.add_widget(self.my_button_motif2)
        self.my_gridlayout3ba1.add_widget(self.my_gridlayout3bawww)
        self.my_gridlayout3ba1.add_widget(self.spinner_child)

        self.my_gridlayout3b1.add_widget(self.my_gridlayout3ba1)

        self.my_gridlayout3ba1 = GridLayout(cols=2, spacing=0, padding=0, size_hint=(0.25, 1.5))

        self.my_gridlayout3baccc = GridLayout(cols=1, spacing=0, padding=0, size_hint=(None, 1), width=150)       
        self.my_button_motif3 = Button(text="Réf 3", font_size=19, background_color=(0.43,0.88,1.2,1))      
        self.my_button_motif3.bind(on_press=self.reset_spinner_grandchild)

        
        # Spinner petit-enfant
       
        self.spinner_grandchild = Spinner(
            text='Sous version document',
            values=[], option_cls=self.custom_spinner_option, color='#00000',background_color=(255,255,255,1), font_size=19
        )
        # Associez les Spinners aux listes d'options
        self.spinner_parent.values = categories
        self.spinner_parent.bind(text=self.update_spinner_child)

        self.spinner_child.bind(text=self.update_spinner_grandchild)



        self.my_gridlayout3baccc.add_widget(self.my_button_motif3)
        self.my_gridlayout3ba1.add_widget(self.my_gridlayout3baccc)
        self.my_gridlayout3ba1.add_widget(self.spinner_grandchild)

        self.my_gridlayout3b1.add_widget(self.my_gridlayout3ba1)

        self.my_gridlayout3ba1 = GridLayout(cols=2, spacing=0, padding=0, size_hint=(0.25, 4))   
        
        self.my_gridlayout3baccc = GridLayout(cols=1, spacing=0, padding=0, size_hint=(None, 1), width=200)        
        self.my_button_commentaire = Button(text="Commentaire", font_size=19, background_color=(0.43,0.88,1.2,1))

        #self.my_button_commentaire.bind(on_press=self.enregistrer_vocale_commentaire)

        self.my_text_input_14 = TextInput(multiline=True, font_size=19)
        self.my_gridlayout3baccc.add_widget(self.my_button_commentaire)
        self.my_gridlayout3ba1.add_widget(self.my_gridlayout3baccc)
        self.my_gridlayout3ba1.add_widget(self.my_text_input_14)
        self.text_inputs.append(self.my_text_input_14)

        self.my_gridlayout3b1.add_widget(self.my_gridlayout3ba1)

        self.my_gridlayout3b.add_widget(self.my_gridlayout3b1)



        #-------------------------------------------------------------------------------

        self.add_widget(self.my_gridlayout3b)

        #-----------------------------------------------------------------------------------------------------------------------------------------


        #2222AAAAAAAAA-----------------------------------------------------------------------------------------------------------------------------------

            # Ajoute un bouton
            

        self.my_gridlayout3b = GridLayout(cols=1, spacing=10, padding=5, size_hint=(1, 3))






        #-------------------------------------------------------------------------------
        
        self.my_gridlayout3b1 = GridLayout(rows=3, spacing=0, padding=0, size_hint=(0.25, 1.5))

        #------------------------------------

        self.my_gridlayout3ba = GridLayout(cols=1, spacing=0, padding=0, size_hint=(0.25, 1))        
        self.my_button_envoyer3b = Button(text="Prestations concernées", font_size=19, background_color=(0.1,0.56,0.9,1))
        self.my_gridlayout3ba.add_widget(self.my_button_envoyer3b)
               

        self.my_gridlayout3b1.add_widget(self.my_gridlayout3ba)

        #------------------------------------

        self.my_gridlayout3ba = GridLayout(cols=1, spacing=0, padding=10, size_hint=(0.25, 4))   
        
        # Créez un ScrollView
        self.scroll_view = ScrollView()

        # Créez un GridLayout pour contenir les boutons
        self.grid_layout = GridLayout(cols=1, spacing=0, size_hint_y=None)

                         
        # Liste d'éléments à afficher dans les boutons
        self.elements = [" Demande générique", "ACFP", "ACTP", "AEEH et compléments", "Aide humaine élèves handicapés",
                          "Allocation adulte handicapé", "Aménagement examens", "Assurance vieillesse des parents au foyer", "CMI Priorité/Invalidité", "CMI Stationnement",
                          "Complément de ressources", "CRP/CPO/UEROS", "Élément 2", "Demandes relative à la vie scolaire", "Dispositif Emploi accompagné",
                          "ESAT", "Formation et insertion professionnelle", "Maintien en Etablissement Amendement Creton", "Marché du travail", "Matériel pédagogique adapté",
                          "Orientation ESMS Adulte", "Orientation ESMS Enfant", "Orientation scolaire", "PCH 1 - Aide humaine à domicile", "PCH 1 - Aide humaine en établissement",
                          "PCH 2 - Aides techniques", "PCH 3 - Aménagement du logement et du véhicule", "PCH 4 - Charges spécifiques et exceptionnelles", "PCH 5 - Aide animalière", "RQTH",
                          "Transport scolaire", "UVPHA", "Autres"]

        
        self.buttons = {}  # Dictionnaire pour stocker les boutons
        self.button_states = {}  # Dictionnaire pour stocker l'état des boutons
        self.list_button_true = []

        for element in self.elements:
            self.button = Button(text=element,
                size_hint_y=None,
                height=25,  # Ajustez la hauteur selon vos besoins
                font_size=19,
                background_color = "#b0b0b0"  # Ajustez la taille de la police ici
                )
            self.button.bind(on_press=self.on_button_press)  # Associez une fonction de gestionnaire d'événements
            self.buttons[element] = self.button
            self.button_states[element] = False  # Initialisez l'état du bouton comme désactivé
            self.grid_layout.add_widget(self.button)  # Ajoutez le bouton à la mise en page
            
        
        # Indiquez au GridLayout de calculer sa taille en fonction de son contenu
        self.grid_layout.bind(minimum_height=self.grid_layout.setter('height'))

        # Ajoutez le GridLayout au ScrollView
        self.scroll_view.add_widget(self.grid_layout)


        # Ajoutez ScrollView à votre interface


        self.my_gridlayout3ba.add_widget(self.scroll_view )
 



               

        self.my_gridlayout3b1.add_widget(self.my_gridlayout3ba)

        #------------------------------------

        self.my_gridlayout3ba = GridLayout(cols=1, spacing=0, padding=0, size_hint=(0.25, 1))        
        #self.my_button_envoyer3b = Button(text=f"{[i for i in self.buttons]}", font_size=19, background_color=(0.1,0.56,0.9,1))
        self.my_button_prestations = Button(text=str(self.list_button_true), font_size=19, background_color=(0.1,0.56,0.9,1))

        self.my_gridlayout3ba.add_widget(self.my_button_prestations)

        self.my_gridlayout3b1.add_widget(self.my_gridlayout3ba)

        #------------------------------------
        
        self.my_gridlayout3b.add_widget(self.my_gridlayout3b1)

        self.add_widget(self.my_gridlayout3b)

        #-----------------------------------------------------------------------------------------------------------------------------------------

        #3333-----------------------------------------------------------------------------------------------------------------------------------

            # Ajoute un bouton
            

        self.my_gridlayout3b = GridLayout(cols=2, spacing=10, padding=5, size_hint=(1, 2))



        self.my_gridlayout3b1 = GridLayout(rows=3, spacing=0, padding=0, size_hint=(1, 1))

        self.my_gridlayout3ba = GridLayout(cols=1, spacing=0, padding=0, size_hint=(1, 2))        
        self.my_button_envoyer3b2 = Button(text="Transmission de la demande au : ", font_size=19, background_color=(0.1,0.56,0.9,1))
        self.my_gridlayout3ba.add_widget(self.my_button_envoyer3b2)

        self.my_gridlayout3b1.add_widget(self.my_gridlayout3ba)

        self.my_gridlayout3ba1 = GridLayout(cols=2, spacing=0, padding=0, size_hint=(1, 2))   

        self.my_gridlayout3bavvv = GridLayout(cols=1, spacing=0, padding=0, size_hint=(None, 1), width=150)        
        self.my_button_service = Button(text="Pôle", font_size=19, background_color=(0.1,0.56,0.9,1))


        self.options = ['JURIDIQUE', 'EVALUATION', 'CMI', 'ESMS', 'PCH', 'PILOTAGE', 'QUALITE', 'DIRECTION',
                         'TRANSFERT']
        
        
        self.my_spinner_service = Spinner(
            text='Sélectionnez un pôle',  # Texte initial
            values=self.options, option_cls=self.CustomSpinnerOption, color='#00000',background_color=(255,255,255,1), font_size=19)  # Liste d'options
            #size_hint=(None, None))
            #size=(150, 44))  # Taille du menu déroulant)
        def on_spinner_select(instance, value):
            # La fonction est appelée lorsque l'utilisateur sélectionne une option
            print(f'Option sélectionnée : {value}')

        self.my_spinner_service.bind(text=on_spinner_select)


        self.my_gridlayout3bavvv.add_widget(self.my_button_service)
        self.my_gridlayout3ba1.add_widget(self.my_gridlayout3bavvv)
        self.my_gridlayout3ba1.add_widget(self.my_spinner_service)

        self.my_gridlayout3b1.add_widget(self.my_gridlayout3ba1)

        self.my_gridlayout3ba1 = GridLayout(cols=2, spacing=0, padding=0, size_hint=(0.25,2))       
        
        self.my_gridlayout3babbb = GridLayout(cols=1, spacing=0, padding=0, size_hint=(None, 1), width=150)   
        self.my_button_envoyer3b1 = Button(text="Personne", font_size=19, background_color=(0.1,0.56,0.9,1))
        self.my_text_input_15 = TextInput(multiline=True, font_size=19)
        self.my_gridlayout3babbb.add_widget(self.my_button_envoyer3b1)
        self.my_gridlayout3ba1.add_widget(self.my_gridlayout3babbb)
        self.my_gridlayout3ba1.add_widget(self.my_text_input_15)
        self.text_inputs.append(self.my_text_input_15)

        self.my_gridlayout3b1.add_widget(self.my_gridlayout3ba1)


        self.my_gridlayout3b.add_widget(self.my_gridlayout3b1)


        #-------------------------------------------------------------------------------


        self.add_widget(self.my_gridlayout3b)
        

        #-----------------------------------------------------------------------------------------------------------------------------------------
  
        #3333aaaaaaaaaaaaaaaa-----------------------------------------------------------------------------------------------------------------------------------

            # Ajoute un bouton
            

        self.my_gridlayout3b = GridLayout(cols=1, spacing=10, padding=5, size_hint=(1, 1))



        #-------------------------------------------------------------------------------
        
        self.my_gridlayout3b1 = GridLayout(rows=1, spacing=0, padding=0, size_hint=(0.25, 0.25))

        #------------------------------------
        self.my_gridlayout3ba = GridLayout(cols=3, spacing=10, padding=0, size_hint=(0.25, 1))        
        self.my_button_enregister = Button(text="Enregistrer", font_size=19, background_color=(0.43,0.88,1.2,1))
        self.my_button_enregister.bind(on_press=self.save_text)
        self.my_button_capture = Button(text="Capturer", font_size=19, background_color=(0.43,0.88,1.2,1))
        self.my_gridlayout3ba.add_widget(self.my_button_enregister)
        self.my_button_capture.bind(on_release=self.capture_and_save_screen)
        self.my_gridlayout3ba.add_widget(self.my_button_capture)
               
        self.my_button_quitter = Button(text="Quitter", font_size=19, background_color=(0.43,0.88,1.2,1))
        
        # Associez la fonction pour quitter l'application au bouton
        self.my_button_quitter.bind(on_release=self.quit_app)

        self.my_gridlayout3ba.add_widget(self.my_button_quitter)

        self.my_gridlayout3b1.add_widget(self.my_gridlayout3ba)

        
        #------------------------------------

        self.my_gridlayout3b.add_widget(self.my_gridlayout3b1)

        self.add_widget(self.my_gridlayout3b)
        
        # Créez un dictionnaire pour stocker les TextInput
        self.text_inputs2 = {
            "Nom": self.my_text_input_1,
            "Prénom": self.my_text_input_2,
            "Commentaire": self.my_text_input_14,
            "Nom service": self.my_text_input_15,
            "Type": self.my_spinner_type,
            "Territoire": self.my_spinner_territoire,
            "Motif1": self.spinner_parent,
            "Motif2": self.spinner_child,
            "Motif3": self.spinner_grandchild,
            "Service": self.my_spinner_service,
            "Date": date_du_jour(),
            "Heure": heure_du_jour(),
            "Agent": self.my_spinner_agent,
            "Cellule": self.my_spinner_cellule,
            "Prestations": self.list_button_true
        }

        # Créez un dictionnaire pour stocker les textes
        self.saved_texts = {}

        #-----------------------------------------------------------------------------------------------------------------------------------------
  

    def save_text(self, instance):
        if self.my_spinner_agent.text != 'Nom Agent':
            if self.my_spinner_cellule.text != 'Cellule':
                if self.my_spinner_type.text != 'Sélectionnez un type':
                    if self.my_text_input_1.text != "":
                        if self.my_text_input_2.text != "":
                            if self.my_spinner_territoire.text != 'Sélectionnez un sujet':
                                if self.spinner_parent.text != 'Sélectionnez le motif 1':
                                    if self.my_spinner_service.text != 'Sélectionnez un pôle':
                                        if len(self.list_button_true) > 0:
                                            # Incrémente la valeur
                                            self.compteur += 1

                                            # Mettez à jour le champ texte avec la nouvelle valeur
                                            self.my_gridlayout_compteur.text = f"{self.compteur} tache(s)"

                                            # Enregistrez le contenu de chaque TextInput dans le dictionnaire saved_texts
                                            for input_name, text_input in self.text_inputs2.items():
                                                try:
                                                    self.saved_texts[input_name] = text_input.text
                                                except:
                                                    self.saved_texts[input_name] = text_input
                                            print(self.saved_texts)


                                            """df = pd.DataFrame(list(zip(self.my_spinner_type, self.my_text_input_1, self.my_text_input_2, self.my_text_input_3, self.my_text_input_3a, self.my_spinner_AdEf, 
                                                                    self.my_text_input_5,self.my_text_input_6, self.my_text_input_7, self.my_text_input_8, self.my_text_input_9, self.my_text_input_10,
                                                                        self.spinner_parent, self.spinner_child, self.spinner_grandchild, self.my_text_input_14, self.my_spinner_service, self.my_text_input_15)),
                                                                    columns =['Type', 'Nom', 'Prénom', 'Nom1', 'Prénom2','Catégorie' ,
                                                                            'DdN', 'Adresse', 'CP', 'Ville', 'Téléphone', 'Date', 'Référence',
                                                                            'Motif1', 'Motif2', 'Motif3', 'Commentaire', 'Service', 'Nom service'])"""
                                            
                                            # Nom du fichier Excel existant
                                            file_name = "Liste_taches.xlsx"

                                            # Chargez le classeur Excel existant
                                            try:
                                                wb = load_workbook(file_name)
                                            except FileNotFoundError:
                                                # Si le fichier n'existe pas, créez un nouveau classeur
                                                wb = Workbook()

                                            # Sélectionnez la feuille de calcul existante ou créez-en une nouvelle
                                            if "Liste_taches" in wb.sheetnames:
                                                ws = wb["Liste_taches"]
                                            else:
                                                ws = wb.create_sheet(title="Liste_tachels")
                                            # Obtenez la prochaine ligne vide dans la colonne A
                                            next_row = ws.max_row + 1

                                            # Insérez la nouvelle date dans la colonne A à la prochaine ligne vide

                                                # Insérez la date dans la cellule A1 de la feuille de calcul
                                            ws[f'A{next_row}'] = self.my_spinner_type.text
                                            ws[f'B{next_row}'] = self.my_text_input_1.text.strip()
                                            ws[f'C{next_row}'] = self.my_text_input_2.text.strip()
                                            ws[f'D{next_row}'] = self.spinner_parent.text
                                            ws[f'E{next_row}'] = self.spinner_child.text
                                            ws[f'F{next_row}'] = self.spinner_grandchild.text
                                            ws[f'G{next_row}'] = self.my_text_input_14.text.strip()
                                            ws[f'H{next_row}'] = self.my_spinner_service.text
                                            ws[f'I{next_row}'] = self.my_text_input_15.text.strip()
                                            ws[f'J{next_row}'] = date_du_jour()
                                            ws[f'K{next_row}'] = heure_du_jour()
                                            ws[f'L{next_row}'] = self.my_spinner_agent.text
                                            ws[f'M{next_row}'] = self.my_spinner_cellule.text
                                            ws[f'N{next_row}'] = ", ".join(self.list_button_true) 
                                            
                                            # Enregistrez le classeur Excel
                                            wb.save(file_name)

                                            # Appelez la fonction pour afficher le Popup
                                            self.show_save_success_popup()

                                            # Réinitialisez tous les widgets
                                            self.reset_widgets()

                                            # Actualisez la date et l'heure
                                            self.my_button_date.text = date_du_jour()
                                            self.my_button_heure.text = heure_du_jour()
                                        else:
                                            self.my_button_prestations.background_color=(1,0,0,1)
                                    else:
                                        self.my_button_service.background_color=(1,0,0,1)
                                else:
                                    self.my_button_motif1.background_color=(1,0,0,1)
                            else:
                                self.my_button_territoire.background_color=(1,0,0,1)
                        else:
                            self.my_text_input_2.background_color=(1,0,0,1)

                    else:
                        self.my_text_input_1.background_color=(1,0,0,1)
                else:
                    self.my_button_type.background_color=(1,0,0,1)
            else:
                self.my_spinner_cellule.background_color=(1,0,0,1)
        else:
            self.my_spinner_agent.background_color=(1,0,0,1)

    def on_button_press(self, button):
        element = button.text
        if self.button_states[element]:
            self.button_states[element] = False
            button.background_color = "#b0b0b0"
            self.list_button_true.remove(element)  # Retirez l'élément de la liste
        else:
            self.button_states[element] = True
            self.list_button_true.append(element)
            button.background_color = (0.1,0.56,0.9,1)

        # Mettez à jour le texte du bouton "Prestations concernées"
        self.my_button_prestations.text = ", ".join(self.list_button_true)


    def quit_app(self, instance):
        App.get_running_app().stop()
 
  
    def update_spinner_child(self, instance, value):
        if value == 'Référenciel Fonctionnel':
            self.spinner_child.values = ["2.1 v1", "2.1 v2","2.2 v1","2.2 v2",]
        elif value == "DEMANDE D'INFORMATIONS DIVERSES":
            self.spinner_child.values = ['Attestation ( retraite, TI…)', 'Coordonnées usagers, partenaires…', 'MDPH et ses prestations',
                                         "Modalité de mise en œuvre d'une décision ( paiement AAH, attribution MPA…)",
                                         "Prestations hors MDPH ( aide ménagère/APA…) Ne concerne pas la MDPH", "Horaires et coordonnées MDPH."]
        elif value == 'EXPLICATION DE DECISION':
            self.spinner_child.values = ['Accord', 'Rejet']
        elif value == "DEMANDE D'ENVOI DE DOCUMENTS DIVERS":
            self.spinner_child.values = ['Ajout de demande', 'Copie cerfa CAF', "Copie du dossier de l'usager", "Envoi d'un dossier MDPH (y compris aménagement d'examens)"]
        elif value == 'RECTIFICATIF DE DECISION':
            self.spinner_child.values = ['Catégorie complément', 'Date', 'Nom', 'Organisme Payeur', "Taux d'incapacité", "Type d'établissement"]
        elif value == 'DEPOT DE DOCUMENT':
            self.spinner_child.values = ['Dossier', 'Pièce', "Pièce non conforme"]
        elif value == 'DUPLICATA DE NOTIFICATION':
            self.spinner_child.values = ['Non reçue', 'Perdue']
        elif value == 'SITUATION URGENTE':
            self.spinner_child.values = ['Envoi vers cellule pilotage', 'Cellule alarmante']
        elif value == 'RECOURS':
            self.spinner_child.values = ['Conciliation', 'RAPO', 'Contentieux']
        elif value == 'TRANSFERT DE DOSSIER':
            self.spinner_child.values = ['Entrant', 'Sortant']
        elif value == 'DEMANDE DE RENDEZ VOUS':
            self.spinner_child.values = [""]
        elif value == 'RENVOI AUX SERVICES':
            self.spinner_child.values = ['Mission Accueil ', 'Mission Accompagnement', 'Mission Appui et Ressources/Logistique', 'Mission EMS', 'Mission Evaluation', 'Mission Numérisation', 'Mission PCH', 'Mission Qualité ', 'Mission Traitement Administratif', 'Boite mail CMI', 'Secrétariat Médical', "Direction"]
        elif value == 'COMMUNICATION INTERROMPUE':
            self.spinner_child.values = [""]
        elif value == 'PORTAIL USAGER':
            self.spinner_child.values = ["Comment déposer une demande en ligne", "Problème de connexion", "Problème d’identifiant", "Problèmes rencontrés sur le portail"]
        elif value == 'MESURE DE SATISFACTION DES USAGERS':
            self.spinner_child.values = ['Satisfait', 'Insatisfait']
        elif value == "AIDE AU REMPLISSAGE D'UN DOSSIER":
            self.spinner_child.values = ['Par téléphone', "A l'accueil physique"]
        elif value == 'PARTENAIRES':
            self.spinner_child.values = ['Référenciel Fonctionnel', 'Demande en ligne', 'Rupture de droit', 'Renseignements sur les Prestations', 'Explications des Décisions', "Modalités de Mise en Œuvre d'une Décision", 'Rectificatif de décision', 'Duplicata de notifications', "Recours : état d'avancement", 'Transfert de dossier', 'Autres (Précisez)']
        else:
            self.spinner_child.values = []

    def update_spinner_grandchild(self, instance, value):
        if value == "MDPH et ses prestations":
            self.spinner_grandchild.values = ['Parcours de scolarisation', 'Parcours professionnel', 'Prestation financière adulte', 'Prestation financière enfant', 'Vie quotidienne ( CMI, SAVS…)']
        elif value == "Modalité de mise en œuvre d'une décision ( paiement AAH, attribution MPA…)":
            self.spinner_grandchild.values = ["Parcours de scolarisation", 'Parcours professionnel', 'Prestation financière adulte', 'Prestation financière enfant', 'Vie quotidienne ( CMI, SAVS…)']
        elif value == 'Accord':
            self.spinner_grandchild.values = ["Parcours de scolarisation", 'Parcours professionnel', 'Prestation financière adulte', 'Prestation financière enfant', 'Vie quotidienne ( CMI, SAVS…)']
        elif value == 'Rejet':
            self.spinner_grandchild.values = ["Parcours de scolarisation", 'Parcours professionnel', 'Prestation financière adulte', 'Prestation financière enfant', 'Vie quotidienne ( CMI, SAVS…)']
        elif value == 'Dossier':
            self.spinner_grandchild.values = ["Adulte", 'Enfant']
        elif value == 'Pièce':
            self.spinner_grandchild.values = ["Pièce administrative", 'Pièce médicale']
        elif value == 'Non reçue':
            self.spinner_grandchild.values = ["CMI-I/P, CMI-S", 'Parcours de scolarisation', 'Parcours professionnel', 'Prestation financière adulte', 'Prestation financière enfant', 'Vie quotidienne ( EMS, SAVS…)']
        elif value == 'Perdue':
            self.spinner_grandchild.values = ["CMI-I/P, CMI-S", 'Parcours de scolarisation', 'Parcours professionnel', 'Prestation financière adulte', 'Prestation financière enfant', 'Vie quotidienne ( EMS, SAVS…)']
        elif value == 'Conciliation':
            self.spinner_grandchild.values = ["Dépôt document complémentaire", "Dépôt d'une demandeé", "Etat d'avancement", 'Modalités de recours']
        elif value == 'RAPO':
            self.spinner_grandchild.values = ["Dépôt document complémentaire", "Dépôt d'une demande", "Etat d'avancement", 'Modalités de recours']
        elif value == 'Contentieux':
            self.spinner_grandchild.values = ["Retour pièces Tribunal", 'Convocation', "mise en œuvre d’un jugement."]
        elif value == 'Boite mail CMI':
            self.spinner_grandchild.values = ["Demande appel photo", 'Duplicata CMI', 'Identifiant/MDP perdus']
        elif value == 'Attestation ( retraite, TI…)':
            self.spinner_grandchild.values = ["Perdu", 'Abimé']
        else:
            self.spinner_grandchild.values = []

            
    def reset_spinner_parent(self, instance):
        # Réinitialiser le Spinner parent à sa valeur par défaut
        self.spinner_parent.text = ''

    def reset_spinner_child(self, instance):
        # Réinitialiser le Spinner enfant à sa valeur par défaut
        self.spinner_child.text = ''

    def reset_spinner_grandchild(self, instance):
        # Réinitialiser le Spinner petit-enfant à sa valeur par défaut
        self.spinner_grandchild.text = ''

    def reset_heure_debut(self, instance):
        # Réinitialiser le Spinner petit-enfant à sa valeur par défaut
        self.my_button_heure_debut_value.text = heure_du_jour()
        
    def reset_heure_fin(self, instance):
        # Réinitialiser le Spinner petit-enfant à sa valeur par défaut
        self.my_button_heure_fin_value.text = heure_du_jour()

    #def enregistrer_vocale_nom(self, instance):
        # Réinitialiser le Spinner petit-enfant à sa valeur par défaut
        #self.my_text_input_1.text = enregistrer_vocale(self.my_button_nom.text)


    #def enregistrer_vocale_commentaire(self, instance):
        # Réinitialiser le Spinner petit-enfant à sa valeur par défaut
        #self.my_text_input_14.text = enregistrer_vocale(self.my_button_commentaire.text)

    def on_key_down(self, instance, keyboard, keycode, text, modifiers):
        # Vérifiez si la touche "Tab" a été enfoncée
        if keycode == 43:  # 9 est le code de la touche "Tab"
            # Trouvez le TextInput actuellement en focus
            current_textinput = None
            for text_input in self.text_inputs:
                if text_input.focus:
                    current_textinput = text_input
                    break

            if current_textinput:
                # Déterminez l'index du TextInput actuel
                current_index = self.text_inputs.index(current_textinput)

                # Passez au TextInput suivant (boucle vers le début si c'est le dernier)
                next_index = (current_index + 1) % len(self.text_inputs)
                self.text_inputs[next_index].focus = True


    def show_save_success_popup(self):
        content = BoxLayout(orientation="vertical")
        message = Label(text="L'enregistrement a été effectué avec succès.")
        close_button = Button(text="Fermer")

        # Créez un Popup avec le contenu
        popup = Popup(title="Enregistrement réussi", content=content, size_hint=(None, None), size=(400, 180), auto_dismiss=False)

        # Associez le bouton "Fermer" pour fermer le Popup
        close_button.bind(on_release=popup.dismiss)

        # Ajoutez le contenu au Popup
        content.add_widget(message)
        content.add_widget(close_button)

        # Ouvrez le Popup
        popup.open()

    def show_capture_success_popup(self):
        content = BoxLayout(orientation="vertical")
        message = Label(text="La capture a été effectué avec succès.")
        close_button = Button(text="Fermer")

        # Créez un Popup avec le contenu
        popup = Popup(title="Capture réussi", content=content, size_hint=(None, None), size=(400, 180), auto_dismiss=False)

        # Associez le bouton "Fermer" pour fermer le Popup
        close_button.bind(on_release=popup.dismiss)

        # Ajoutez le contenu au Popup
        content.add_widget(message)
        content.add_widget(close_button)

        # Ouvrez le Popup
        popup.open()


    class custom_spinner_option(SpinnerOption):
        # Personnalisez la couleur du texte pour les options ici
        background_normal=''
        background_color = (0, 0, 0, 1)  # Couleur de fond (blanc)
        color = (1, 1, 1, 1)  # Couleur du texte (noir)

    class CustomSpinnerOption(SpinnerOption):
        background_normal = ''  
        background_color = (0, 0, 0, 1)  # Couleur de fond (blanc)
        color = (1, 1, 1, 1)  # Couleur du texte (noir)

    def reset_widgets(self):
        # Réinitialisez les TextInputs en vidant leur texte
        for text_input in self.text_inputs:
            text_input.text = ''
            text_input.background_color=(1,1,1,1)

        # Réinitialisez les Spinners en rétablissant leur valeur par défaut
        
        self.my_spinner_territoire.text ='Sélectionnez un sujet'
        self.spinner_parent.text ='Document'
        self.spinner_child.text ='Version document'
        self.spinner_grandchild.text ='Sous version document'
        self.my_spinner_service.text ='Sélectionnez un pôle'
        self.my_spinner_type.text = "Sélectionnez un type"

        self.my_button_type.background_color=(0.1,0.56,0.9,1)
        self.my_button_service.background_color=(0.1,0.56,0.9,1)
        self.my_button_territoire.background_color=(0.1,0.56,0.9,1)
        self.my_button_motif1.background_color=(0.43,0.88,1.2,1)
        self.my_spinner_agent.background_color=(0,1,1,1)
        self.my_spinner_cellule.background_color=(1,1,1,0.3)
        self.my_button_prestations.background_color=(0.1,0.56,0.9,1)

        for element, button in self.buttons.items():
            self.button_states[element] = False
            button.background_color = "#b0b0b0"
            self.list_button_true = []

    def capture_and_save_screen(self, instance):
            # Capturez un écran de l'application
            im = ImageGrab.grab()  # Utilisez la fonction grab() de Pillow

            # Définissez le chemin du dossier de sauvegarde
            folder_path = "captures_ecran"

            # Assurez-vous que le dossier existe, sinon créez-le
            if not os.path.exists(folder_path):
                os.makedirs(folder_path)

            # Définissez le chemin complet du fichier de capture
            file_path = os.path.join(folder_path, f"{str(date_du_jour()).replace('/','_')} {str(heure_du_jour()).replace(':','_')} {str(self.my_spinner_agent.text)}.png")

            # Enregistrez l'image capturée dans le dossier spécifié
            im.save(file_path, "PNG")
            
            # Appeleaz la fonction pour afficher le Popup
            self.show_capture_success_popup()
            

        #-----------------------------------------------------------------------------------------------------------------------------------------

          

class MyApp(App):
    def build(self):
        return MyBoxLayout()

if __name__ == '__main__':
    MyApp().run()

